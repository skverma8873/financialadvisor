import os, re, uuid, pyodbc, csv
from agents.agent import Agent
from prompt_templates import mf_product_template
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage
from termcolor import colored
from decimal import Decimal
import simplejson as json
from datetime import date, datetime
from helper import Helper, MessageType


import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from decimal import Decimal
from datetime import datetime

class MF_Products_Agent(Agent):

    def invoke(self, question):

        agent_name = "MF PRODUCTS AGENT"
        Helper.print(MessageType.AGENT_MESSAGE, "AGENT NAME", agent_name)

        json_object = self.product_logic(question)

        # Append as AIMessage is an object.
        #aimessage_object = AIMessage(content = json.dumps(json_object))
        aimessage_object = {"role" : "assistant", "content" : str(json_object["response"])}
        self.messages.append(aimessage_object)

        # Update state object.
        self.update_state("next_agent", json_object["next_agent"])
        self.update_state("previous_agents_response", json_object["response"])
        self.update_state("mf_product_response", aimessage_object)

        Helper.print(MessageType.AI_MESSAGE, "AI RESPONSE", aimessage_object)

        self.updateflow(agent_name)
        
        return self.state
        
    def product_logic(self, question):

        #GENERATE QUERY
        system_prompt_query = mf_product_template.template2.format(
            Placeholder1 = mf_product_template.reference_table_schema,
            Placeholder2 = mf_product_template.reference_column_details)
        
        message_query = [
            {"role" : "system", "content" : system_prompt_query},
            {"role" : "user", "content" : f"User Question : {question}"}
        ]

        Helper.print(MessageType.PROMPT_MESSAGE, "PROMPT FOR SQL QUERY", message_query)

        response = self.chatllm.invoke(message_query) # response is in string format (but valid json inside).
        query_json = json.loads(response.content) # converts to json.
        sql_query = MF_Product_Tools.clean_query(query_json["sql_query"]) # fetch sql_query from it.
        
        Helper.print(MessageType.TEXT_LOG, "SQL QUERY", sql_query)

        #GENERATE DATA FROM QUERY
        data = MF_Product_Tools.fetch_data_from_sql(sql_query)
        Helper.print(MessageType.TEXT_LOG, "DB RECORDS", data)


        dataset_string = ""
        for index, item in enumerate(data,1):
            dataset_string += f"""
        
            Database Resultset {index}:
            {json.dumps(item[:5], default = MF_Product_Tools.custom_serializer)}

            """
        dataset_string = " ".join(dataset_string.strip().split())


        #GENERATE INSIGHITS RESPONSE FROM DATA
        r_response = "EMPTY"
        if(len(self.state["reviewer_response"]) > 0):
            r_response = self.state["reviewer_response"][-1].content

        system_prompt_insights = mf_product_template.template3.format(
            Placeholder1 = dataset_string,
            Placeholder2 = r_response)

        message_insights = [
            {"role" : "system", "content" : system_prompt_insights},
            {"role" : "user", "content" : f"** User Question ** : {question}"}
        ]

        Helper.print(MessageType.PROMPT_MESSAGE, "PROMPT FOR INSIGHTS", message_insights)

        response = self.chatllm.invoke(message_insights) # response is returned in string (but valid json)
        json_object = json.loads(response.content) # convert string to json

        #GENERATE CSV FILE WITH DATA
        retVal = ""

        if data:
            file_name = f"{str(uuid.uuid4())}.xlsx"
            MF_Product_Tools.write_multiple_sheets_to_excel(data, file_name)
            anchor_tag = f'<a href = "http://localhost:5000/download/{file_name}">Download</a>'
            retVal = {"next_agent" : "reviewer",  "response" : json_object["response"] + "\n\n" + f"Reference File : {anchor_tag}" }
            
        else:
            retVal = {"next_agent" : "reviewer",  "response" : json_object["response"]}
        
        return retVal
    
class MF_Product_Tools():

    def clean_query(query):
        query = re.sub(r"```sql\n|\n```", "", query).strip()
        query = re.sub(r"\n", "", query).strip()
        return query

    def fetch_data_from_sql(query):
        """
        Connects to an MS SQL database, executes a query, and returns the data as a list of dictionaries.

        :param query: SQL query to execute
        :param connection_string: Connection string for the database
        :return: List of dictionaries containing query results
        """
        try:
            connection_string = (
                'Driver={ODBC Driver 17 for SQL Server};'
                'Server=DESKTOP-RQN38R0;'
                'Database=Mutual_Funds;'
                'UID=sa;'
                'PWD=mssql@123;'
            )
            
            # Connect to the database
            connection = pyodbc.connect(connection_string)
            cursor = connection.cursor()

            # Execute the query
            cursor.execute(query)

            all_results = []  # To store all result sets

            # Process each result set
            while True:
                # Fetch all rows from the current result set
                columns = [column[0] for column in cursor.description]  # Get column names
                rows = cursor.fetchall()  # Fetch all data

                # Dynamically populate data into a list of dictionaries
                if rows:  # Check if rows are not empty
                    result = [dict(zip(columns, row)) for row in rows]
                else:  # If no rows found, append "No Records Found"
                    result = [{"Records from Database": "No details found."}]

                # Append the current result set to the list of all results
                all_results.append(result)

                # Move to the next result set
                if not cursor.nextset():
                    break

            return all_results  # Return all result sets
        except Exception as e:
            print(f"An error occurred: {e}")
            return []

        finally:
            # Close the connection
            cursor.close()
            connection.close()

    def convert_value(value):
        """
        Convert different data types to a format suitable for Excel.
        This ensures that Decimal, datetime, and None types are handled appropriately.
        """
        if isinstance(value, Decimal):
            return float(value)  # Convert Decimal to float
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')  # Convert datetime to string
        elif value is None:
            return ""  # Convert None to empty string
        return value  # Return value as is for other types

    def write_multiple_sheets_to_excel(data, file_name):
        """
        Write multiple lists of data (dictionaries) to one Excel file, simulating multiple sheets.

        :param data: A list of lists, where each inner list contains dictionaries with data for a sheet.
        :param file_path: Path to save the Excel file.
        :param sheet_names: A list of sheet names corresponding to each dataset in data.
        """
        try:
            # Get the current working directory
            cwd = os.getcwd()
            
            # Create the 'assets' folder path
            assets_folder = os.path.join(cwd, '06_chat_using_langgraph','assets')
            
            # Ensure the 'assets' folder exists
            os.makedirs(assets_folder, exist_ok=True)
            
            # Create the full path for the CSV file
            file_path = os.path.join(assets_folder, file_name)

            # Create a workbook
            wb = openpyxl.Workbook()
            
            for index, dataset in enumerate(data, 1):

                sheet_name = f"Sheet {index}"
                    
                # Create a new sheet or use the existing one
                if index == 0:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)

                for item in dataset:
                    # Check if the dataset is empty
                    if item:
                        # Writing headers (keys of the first dictionary)
                        headers = list(item.keys())  # Convert dict_keys to a list
                        ws.append(headers)  # Append the headers
                        break

                # Write the data rows, converting values
                for row in dataset:
                    converted_row = [MF_Product_Tools.convert_value(row[key]) for key in headers]
                    ws.append(converted_row)
                else:
                    # If no data, add a "No Records Found" row
                    ws.append(["No Records Found"])
            

            # Remove the extra default sheet if it still exists
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Save the workbook
            wb.save(file_path)
            print(f"Excel file '{file_path}' has been created successfully with multiple sheets.")
        
        except Exception as e:
            print(f"An error occurred: {e}")

    def custom_serializer(obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()  # Converts date or datetime to ISO 8601 string
        raise TypeError(f"Type {type(obj)} not serializable")

    def write_dict_to_csv(data, file_name):
        """
        Write a list of dictionaries to a CSV file.

        Parameters:
        - data (list of dict): The data to write into the CSV file.
        - file_name (str): Name of the file to save the data.
        """
        if not data:
            print("The data list is empty. No file created.")
            return
        
        # Get the current working directory
        cwd = os.getcwd()
        
        # Create the 'assets' folder path
        assets_folder = os.path.join(cwd, '10_tts_demo','assets')
        
        # Ensure the 'assets' folder exists
        os.makedirs(assets_folder, exist_ok=True)
        
        # Create the full path for the CSV file
        file_path = os.path.join(assets_folder, file_name)

        # Extracting the keys from the first dictionary as the header
        headers = data[0].keys()

        try:
            with open(file_path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=headers)
                
                # Writing the header
                writer.writeheader()
                
                # Writing the data rows
                writer.writerows(data)
            
            print(f"CSV file '{file_name}' has been created successfully.")
        except Exception as e:
            print(f"An error occurred: {e}")
